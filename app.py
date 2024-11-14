from flask import Flask, render_template, request, redirect, url_for, flash
from datetime import datetime
import openpyxl
from openpyxl import Workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key'

questions_file = 'static/files/questions.xlsx'
result_file = 'static/files/submissions.xlsx'

# Load questions from Excel
def load_questions():
    wb = openpyxl.load_workbook(questions_file)
    sheet = wb['question']
    questions = []
    for row in range(2, sheet.max_row + 1):
        label = sheet.cell(row=row, column=3).value
        if label:
            questions.append(label)
    return questions

# Load or create result file
def load_or_create_result_file():
    try:
        wb = openpyxl.load_workbook(result_file)
    except FileNotFoundError:
        wb = Workbook()
        wb.save(result_file)
    return wb

# Save form data
def save_data(date, shift, visitor_data):
    wb = load_or_create_result_file()
    sheet = wb.active
    header = ['Date', 'Shift'] + [f"Visitor Name {i+1}" for i in range(len(visitor_data))]

    # Add header if not present
    if sheet.max_row == 1:
        for col, h in enumerate(header, start=1):
            sheet.cell(row=1, column=col, value=h)

    # Check if date and shift already exist, overwrite the row
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == date and sheet.cell(row=row, column=2).value == shift:
            for col, value in enumerate(visitor_data, start=3):
                sheet.cell(row=row, column=col, value=value)
            wb.save(result_file)
            return

    # If not found, append a new row
    new_row = [date, shift] + visitor_data
    sheet.append(new_row)
    wb.save(result_file)

# Route for displaying the form
@app.route('/', methods=['GET', 'POST'])
def index():
    questions = load_questions()
    if request.method == 'POST':
        date = request.form.get('date')
        shift = request.form.get('shift')
        visitor_data = [request.form.get(f'visitor_{i}') for i in range(len(questions))]

        if not date or not shift:
            flash("Please fill in all fields before submitting.")
            return redirect(url_for('index'))

        # Save the form data
        save_data(date, shift, visitor_data)
        flash("Data submitted successfully!")
        return redirect(url_for('index'))

    return render_template('index.html', questions=questions)

if __name__ == '__main__':
    app.run(debug=True, port=5001)
